#~~~~~~~~~~~~~~~~~~~~~~~~~~~~#
#     CODE DEVELOPED BY      # 
#   Miguel Ant. Linares S.   #
#~~~~~~~~~~~~~~~~~~~~~~~~~~~~#

from io import BytesIO
import pandas as pd
import requests
import openpyxl
import re
import os

# Dictionary where we are going to save data
data = {
    "asof":[],
    "asset_class": [],
    "sub_asset_class": [],
    "type": [],
    "trade_count":[],
    "unique_sec_ids": [],
    "$trades_000s": []
}

# Important routes
# dir = 'C:/Users/mlinares/Documents/GitHub/Web-Scrapping/FINRA_ICE_DATA_SERVICES_WS/response.xlsx'
result_dir = 'C:/Users/mlinares/Documents/GitHub/Web-Scrapping/FINRA_ICE_DATA_SERVICES_WS/finra_ice_results.csv'
url = 'https://cdn.finra.org/trace/FINRA_IDS_STAR.xlsx'

# Download Files from webpage
r = requests.get(url)
# if r.status_code == 200:
#     with open(dir, 'wb') as file:
#         file.write(r.content)

# Open Excel that we downloaded and select the sheet
wb = openpyxl.load_workbook(BytesIO(r.content))
ws = wb['TradingActivity']

# Extract the date out of the document
asof = ws.cell(column=3, row=7).value
asof = asof.strftime('%m/%d/%Y')

header_idx = 0
current_type = None
current_asset = None
current_subasset = None
ct_type = 0
header_text = ['UMBS', 'FNMA', 'FHLMC', 'GNMA', 'OTHER AGENCY', 'ASSET CLASS', 'TRADE', 'COUNT', 'UNIQUE', "SEC ID'S", '$ TRADES', "(000'S)", 'INVESTMENT GRADE', 'NON-INVESTMENT GRADE †']

# Iterate the loop to read the cell values and parsed values
for row in range(8, ws.max_row):
    for col in ws.iter_cols(1, ws.max_column):

        if col[row].value == 'ASSET CLASS':
            header_idx = row

        if col[row].font.bold and col[row].value != None and col[row].value not in header_text:
            current_asset = col[row].value
        
        if col[row].value != None:

            # Regex we use to parse the data
            is_number = re.findall(r'[0-9.\*]+', str(col[row].value))
            is_subasset = re.search(r'([A-Z&\/]+\s?)+([0-9]{0,2}[A-Z])?', str(col[row].value))

            if is_number and not is_subasset and col[row].value not in header_text:
                if ct_type == 3:
                    ct_type = 0
                
                ct_type +=1

                if ct_type == 1:
                    data['trade_count'].append(is_number[0])
                    current_type = col[header_idx].value

                    if 'ASSET CLASS' not in current_type:
                        data['type'].append(current_type)

                if ct_type == 2:
                    data['unique_sec_ids'].append(is_number[0])
                
                if ct_type == 3:
                    data['$trades_000s'].append(is_number[0])
                    if current_asset == current_subasset:
                        data['sub_asset_class'].append(None)
                    else:
                        data['sub_asset_class'].append(current_subasset)
                    data['asset_class'].append(current_asset)
                    data['asof'].append(asof)

            if is_subasset and (col[row].value not in header_text or col[row].value == 'OTHER'):
                current_subasset = is_subasset.group()

# for key in list(data.keys()):
#     print(len(data[key]))

# for key in list(data.keys()):
#     print(data[key])

# Create Dataframe with the data we scrape
df = pd.DataFrame(data)

# Check if another file is saved already
if os.path.isfile(result_dir):
    df_old = pd.read_csv(result_dir)
    asof_old = df_old.iloc[-1:, 0].values
    asof_old = asof_old[0]
    
    #Prevent data duplication by cheching that the asof date is different
    if str(asof) != str(asof_old):
        df = pd.concat([df_old, df], ignore_index=True)
        df.to_csv(result_dir, index=False)
        print('Done...')
    else:
        print('No new data...')
else:
    df.to_csv(result_dir, index=False)
    print('Done...')


# input()